from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from roster import *
from datetime import datetime
import time
from yahoo_oauth import OAuth2
import yahoo_fantasy_api as yfa
import json
import operator
from openpyxl import Workbook, load_workbook
from update_sheets import *
from roster import *

def clear_sheet(wb, sheet):
    workbook = load_workbook(wb)
    sh = workbook[sheet]

    for row in sh['C6:P86']:
        for cell in row:
            cell.value = None

    workbook.save(wb)

def load_position(position, roster, owner):
    positions = { #end row + 1 for inclusivity
        "QB": [6, 16],
        "WR": [17, 34],
        "RB": [35, 47],
        "TE": [49, 58],
        "K": [59, 62],
        "LB": [63, 72],
        "DB": [73, 79],
        "DL": [80, 86]
    }
    wb = load_workbook('dff.xlsx')
    ws = wb['S8 End Roster']
    owner_column = -1
    for col in range (1, 18):
        char = get_column_letter(col)
        if (ws[char + str(2)].value == owner):
            owner_column = col
    
    # Create a list of names to load onto spreadsheet where position matches via list comprehension
    names = [sub.name for sub in roster if sub.position == position]
    
    # Append to spreadsheet at start row
    for row, text in enumerate(names, start=positions[position][0]):
        ws.cell(column=owner_column, row=row, value=text)

    wb.save('dff.xlsx')

def load_team(owner, roster):
    print('DEBUG-' + owner)
    position_types = ["QB", "RB", "WR", "TE", "K", "LB", "DB", "DL"]

    for position in position_types:
        load_position(position, roster, owner)
    
def strip_team_key(team_key):
    return int(team_key[team_key.rindex('.')+1:])

def get_post17_transacts(dynasty, rosters):
    print('N/A')

def get_rosters(dynasty, team_keys, owner_names, week):
    # Reassign week when final transaction week exceeds league end week
    week_param = 1
    if week == 18:
        week_param = 17
    else:
        week_param = week

    
    rosters = []
    team = []

    # Assign position per player 
    for i in range (0,len(team_keys)):
        for item in dynasty.to_team(team_keys[i]).roster(week_param):
            position = ""
            if item["eligible_positions"][0] == "D": 
                position = item["eligible_positions"][1]
            else:
                position = item["eligible_positions"][0]
            

            # player_details_name = ''
            # if '\'' in item["name"]:
            #     player_details_name = item["name"].rsplit('\'', 1)[1]
            #     print(player_details_name + dynasty.player_details(player_details_name)[0]["editorial_team_full_name"])
            # else:
            #     player_details_name = item["name"]

            player = Player(
                item["name"], 
                position, 
                item["player_id"]
                #,dynasty.player_details(player_details_name)[0]["editorial_team_full_name"]
                , " " # Filler
                )
            team.append(player)
        roster = Roster(owner_names[i], team_keys[i], team)
        team = []
        rosters.append(roster)
    
    rosters.sort()

    # Get post league end week transactions
    if week == 18:
        end_dt_as_dt = datetime.strptime(dynasty.settings()['end_date'], '%Y-%m-%d')


        raw_transactions = []
        add_transactions = dynasty.transactions("add", 15)
        raw_transactions.extend(add_transactions)
        drop_transactions = dynasty.transactions("drop", 15)
        
        # Get every unique transaction - Add/drop transactions have some overlap
        for drop in drop_transactions:
            flag = 'n'
            for at in add_transactions:
                if drop['transaction_id'] == at['transaction_id']:
                    flag = 'y'
                    break
            if flag == 'n':
                raw_transactions.extend(drop_transactions)
            flag = 'n'

        # Get transaction only if transaction is past the league end date
        end_transactions = [x for x in raw_transactions if 
                                end_dt_as_dt < datetime.strptime(time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(int(x['timestamp']))), '%Y-%m-%d %H:%M:%S')]

        for x in end_transactions:
            if x["type"] == 'add/drop':
                # player_details_name = ''
                # if '\'' in x["players"]["0"]['player'][0][2]['name']['full']:
                #     player_details_name = x["players"]["0"]['player'][0][2]['name']['full'].rsplit('\'', 1)[1]
                # else:
                #     player_details_name = x["players"]["0"]['player'][0][2]['name']['full']

                # Standardize position before appending
                position = ''
                if x["players"]["0"]['player'][0][4]['display_position'] in ['S','CB']:
                    position = 'DB'
                else:
                    position = x["players"]["0"]['player'][0][4]['display_position']
                added_player = Player(
                    x["players"]["0"]['player'][0][2]['name']['full'], 
                    position, 
                    x["players"]["0"]['player'][0][1]["player_id"]
                    # ,dynasty.player_details(player_details_name)[0]["editorial_team_full_name"]
                    , " "  # Filler
                    )
                roster_index = strip_team_key(x["players"]["0"]['player'][1]['transaction_data'][0]['destination_team_key']) - 1
                roster = rosters[roster_index].roster
                roster.append(added_player)
                
                dropped_player_id = x["players"]["1"]["player"][0][1]["player_id"]
                for player in roster:
                    if int(player.player_id) == int(dropped_player_id):
                        roster.remove(player)
                        break
                
            elif x["type"] == 'add':
                # player_details_name = ''
                # if '\'' in x["players"]["0"]['player'][0][2]['name']['full']:
                #     player_details_name = x["players"]["0"]['player'][0][2]['name']['full'].rsplit('\'', 1)[1]
                # else:
                #     player_details_name = x["players"]["0"]['player'][0][2]['name']['full']
                added_player = Player(
                    x["players"]["0"]['player'][0][2]['name']['full'], 
                    x["players"]["0"]['player'][0][4]['display_position'], 
                    x["players"]["0"]['player'][0][1]["player_id"]
                    # ,dynasty.player_details(player_details_name)[0]["editorial_team_full_name"]
                    , ' ' # Filler
                    )
                
                roster_index = strip_team_key(x["players"]["0"]['player'][1]['transaction_data'][0]['destination_team_key']) - 1
                roster = rosters[roster_index].roster
                roster.append(added_player)

            elif x["type"] == 'drop':
                dropped_player_id = x["players"]["0"]['player'][0][1]["player_id"]

                for player in roster:
                    if int(player.player_id) == int(dropped_player_id):
                        roster.remove(player)
                        break

        
        # Get list of player IDs per roster
        # IDs = [player.player_id for player in roster]
        # dynasty.player_details([player_IDs])

    for roster in rosters:
        # get player IDs for a singular roster
        roster_player_IDs = [int(player.player_id) for player in roster.roster]
        # get a roster's player details
        roster_player_details = dynasty.player_details(roster_player_IDs)

        for i in range(len(roster_player_details)):
            for j in range(len(roster.roster)):
                if int(roster_player_details[i]['player_id']) == roster.roster[j].player_id:
                    roster.roster[j].team = roster_player_details[i]['editorial_team_full_name']
                    break
    return rosters