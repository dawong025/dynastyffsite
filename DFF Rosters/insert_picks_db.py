import sqlite3
from yahoo_oauth import OAuth2
import yahoo_fantasy_api as yfa
from update_sheets import *
from roster import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
sc = OAuth2(None, None, from_file='oauth2.json')
gm = yfa.Game(sc, 'nfl')
import time

dff_db = sqlite3.connect("dynasty-db.db")

cursor = dff_db.cursor()

wb = load_workbook('DFF Rosters/draft.xlsx')
ws = wb['Rookies']

dff_db = sqlite3.connect("dynasty-db.db")
league_years = {
    2018: '380.l.76665',
    2019: '390.l.33135',
    2020: '399.l.20675',
    2021: '406.l.78372',
    2022: '414.l.105716',
    2023: '423.l.202017',
    2024: '449.l.35107'
}
league_year = 2023
league_week = 1

cursor = dff_db.cursor()
leagues = gm.league_ids(year = league_year)
dynasty = gm.to_league(league_years[league_year])

def insert_picks_db(wb, ws):
    for row in ws.iter_rows(min_row = 3, min_col=1, max_row=393, max_col=6):
        if row[0].value:
            year = int(row[0].value)
            pick = str(row[1].value)
            member_name = str(row[2].value)
            player_name = str(row[3].value)
            position = str(row[4].value)
            team = str(row[5].value)
            player_id = -1

            cursor.execute(
                '''
                    SELECT member_id FROM league_members where lower(member_name) = lower(?) and year = ?
                ''', (member_name, year,))
            member_id = cursor.fetchone()[0]

            try:
                cursor.execute(
                    '''
                        SELECT player_id from player where lower(name) = lower(?) and position = ? 
                    ''', (player_name, position)
                )
                player_id = cursor.fetchone()[0]
            except:
                # print(player_name+position+member_name)
                player_id = dynasty.player_details(player_name)[0]['player_id']

            cursor.execute(
                '''
                    INSERT INTO picks (year, pick, member_id, player_id, player_name, team)
                    VALUES(?,?,?,?,?,?)
                ''', (year, pick, member_id, player_id, player_name, team)
            )
        dff_db.commit()



insert_picks_db(wb,ws)





