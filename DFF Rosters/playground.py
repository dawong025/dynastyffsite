import sqlite3
from yahoo_oauth import OAuth2
import yahoo_fantasy_api as yfa
from update_sheets import *
from roster import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
sc = OAuth2(None, None, from_file='oauth2.json')
gm = yfa.Game(sc, 'nfl')
import time

dff_db = sqlite3.connect("dynasty-db.db")
league_years = {
    2018: '380.l.76665',
    2019: '390.l.33135',
    2020: '399.l.20675',
    2021: '406.l.78372',
    2022: '414.l.105716',
    2023: '423.l.202017',
    2024: '449.l.35107'
}
league_year = 2018
league_week = 1

cursor = dff_db.cursor()
leagues = gm.league_ids(year = league_year)
dynasty = gm.to_league(league_years[league_year])


tms = dynasty.teams()

team_keys = []
owner_names = []
league = {}
for key in tms:
    league[dynasty.teams()[key]["managers"][0]["manager"]["nickname"]] = key

wb = load_workbook('DFF Rosters/draft.xlsx')
ws = wb['Rookies']


player_names, yahoo_players = [], []

# for row in ws.iter_rows(min_row = 3, min_col=1, max_row=393, max_col=6):
#     if row[0].value: #not null
#         player_name = str(row[3].value)

#         player_names.append(player_name)

temp_arr = dynasty.player_stats(31135, 'season')
# for i in range(len(temp_arr)): #player_details
#     print(temp_arr[i]['name']['full'] + ' ' + temp_arr[i]['player_id']) 

print(temp_arr[0])



    


            






