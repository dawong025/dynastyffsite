'''
rosters_new.py - Used to get rosters and load them onto sheets
'''

from yahoo_oauth import OAuth2
import yahoo_fantasy_api as yfa
from openpyxl import Workbook, load_workbook
from update_sheets import *
from roster import *
sc = OAuth2(None, None, from_file='oauth2.json')
gm = yfa.Game(sc, 'nfl')
import time

league_years = {
    2018: '380.l.76665',
    2019: '390.l.33135',
    2020: '399.l.20675',
    2021: '406.l.78372',
    2022: '414.l.105716',
    2023: '423.l.202017',
    2024: '449.l.35107'
}
league_year = 2023
league_week = 1

leagues = gm.league_ids(year = league_year)
dynasty = gm.to_league(league_years[league_year])

tms = dynasty.teams()

team_keys = []
owner_names = []
for key in tms:
    team_keys.append(key)
    owner_names.append(dynasty.teams()[key]["managers"][0]["manager"]["nickname"])

league = dict(zip(owner_names, team_keys))

# Testing per roster output
# for roster in rosters:
#     for player in roster.roster:
#         print(player)

#     print('-------------------------------')


wb = load_workbook('DFF Rosters/rosters/dff.xlsx')
ws_template = wb['Roster Template']

'''
Load rosters onto spreadsheet for a given year
'''
# while (league_week <= dynasty.end_week() + 1):
#     print(league_week)

#     rosters = get_rosters(dynasty, team_keys, owner_names, league_week)
#     ws = wb.copy_worksheet(ws_template)
#     ws.title = f'Week {league_week}-{league_year} Roster'

#     for i in range(0, len(owner_names)):
#         load_team(wb, ws, rosters[i].owner_name, rosters[i].roster, league_year) 
#     print(f'{league_week} is done.')
#     league_week += 1

# curr_wb = load_workbook(f'dff - {league_year}.xlsx')
# curr_wb.remove(curr_wb['Roster Template'])

#------------------------------------------------------------  
'''
Load rosters for a given week
'''         
rosters = get_rosters(dynasty, team_keys, owner_names, 1)
ws = wb.copy_worksheet(ws_template)
ws.title = f'Week {league_week}-{league_year} Roster'

for i in range(0, len(owner_names)):
    load_team(wb, ws, rosters[i].owner_name, rosters[i].roster, league_year) 

'''
Save rosters to a new workbook for a given league year
'''
curr_wb = load_workbook(f'DFF Rosters/rosters/dff - {league_year}.xlsx')
curr_wb.remove(curr_wb['Roster Template'])
curr_wb.save(f'DFF Rosters/rosters/dff - {league_year}.xlsx')