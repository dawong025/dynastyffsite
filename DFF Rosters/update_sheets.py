from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from roster import *
from datetime import datetime
import time
from yahoo_oauth import OAuth2
import yahoo_fantasy_api as yfa
from update_sheets import *
from roster import *

# Clears a roster sheet
def clear_sheet(wb, sheet):
    workbook = load_workbook(wb)
    sh = workbook[sheet]

    for row in sh['C6:P86']:
        for cell in row:
            cell.value = None

    workbook.save(wb)

# Loads a position on the sheet based on owner e.g. load all the WRs for Darren
def load_position(wb, ws, position, roster, owner, year):
    # workbook = load_workbook(wb)
    positions = { #end row + 1 for inclusivity
        "QB": [6, 16],
        "WR": [17, 34],
        "RB": [35, 47],
        "TE": [49, 58],
        "K":  [59, 62],
        "LB": [63, 72],
        "DB": [73, 79],
        "DL": [80, 86]
    }

    owner_column = -1
    for col in range (1, 18):
        char = get_column_letter(col)
        if (ws[char + str(2)].value == owner):
            owner_column = col
    
    # Create a list of names to load onto spreadsheet where position matches via list comprehension
    names = [sub.name for sub in roster if sub.position == position]
    
    # Append to spreadsheet at start row
    for row, text in enumerate(names, start=positions[position][0]):
        ws.cell(column=owner_column, row=row, value=text)

    wb.save(f'DFF Rosters/rosters/dff - {year}.xlsx')

# Loads all the positions for an owner via the load position
def load_team(wb, ws, owner, roster,year):
    print('DEBUG-' + owner)
    position_types = ["QB", "RB", "WR", "TE", "K", "LB", "DB", "DL"]

    for position in position_types:
        load_position(wb, ws, position, roster, owner, year)

# Given a sorted list of rosters, strip the team key for easy indexing for each time  
def strip_team_key(team_key):
    return int(team_key[team_key.rindex('.')+1:])

# Get rosters for every team owner at a specified week
def get_rosters(dynasty, team_keys, owner_names, week):
    # Reassign week when final transaction week exceeds league end week
    week_param = 1
    # if week == 18:
    if week == dynasty.end_week() + 1:
        week_param = dynasty.end_week()
    else:
        week_param = week
    
    rosters = []
    team = []

    # Assign position per player 
    for i in range (0,len(team_keys)):
        for item in dynasty.to_team(team_keys[i]).roster(week_param):
            position = ""
            if item["eligible_positions"][0] == "D": 
                position = item["eligible_positions"][1]
            else:
                position = item["eligible_positions"][0]

            player = Player(
                item["name"], 
                position, 
                item["player_id"]
                )
            team.append(player)
        roster = Roster(owner_names[i], team_keys[i], team)
        team = []
        rosters.append(roster)
    
    rosters.sort()

    # Get post league end week transactions
    if week == dynasty.end_week() + 1:
        end_dt_as_dt = datetime.strptime(dynasty.settings()['end_date'], '%Y-%m-%d')

        raw_transactions = []
        add_transactions = dynasty.transactions("add", 15)
        raw_transactions.extend(add_transactions)
        drop_transactions = dynasty.transactions("drop", 15)
        
        # Get every unique transaction - Add/drop transactions have some overlap
        for drop in drop_transactions:
            flag = 'n'
            for at in add_transactions:
                if drop['transaction_id'] == at['transaction_id']:
                    flag = 'y'
                    break
            if flag == 'n':
                raw_transactions.extend(drop_transactions)
            flag = 'n'

        # Get transaction only if transaction is past the league end date
        end_transactions = [x for x in raw_transactions if 
                                end_dt_as_dt < datetime.strptime(time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(int(x['timestamp']))), '%Y-%m-%d %H:%M:%S')]

        for x in end_transactions:
            if x["type"] == 'add/drop':
                # Standardize position before appending
                position = ''
                if x["players"]["0"]['player'][0][4]['display_position'] in ['S','CB']:
                    position = 'DB'
                else:
                    position = x["players"]["0"]['player'][0][4]['display_position']
                added_player = Player(
                    x["players"]["0"]['player'][0][2]['name']['full'], 
                    position, 
                    int(x["players"]["0"]['player'][0][1]["player_id"])
                    )
                roster_index = strip_team_key(x["players"]["0"]['player'][1]['transaction_data'][0]['destination_team_key']) - 1
                roster = rosters[roster_index].roster
                roster.append(added_player)
                
                dropped_player_id = x["players"]["1"]["player"][0][1]["player_id"]
                for player in roster:
                    if int(player.player_id) == int(dropped_player_id):
                        roster.remove(player)
                        break
                
            elif x["type"] == 'add':
                added_player = Player(
                    x["players"]["0"]['player'][0][2]['name']['full'], 
                    x["players"]["0"]['player'][0][4]['display_position'], 
                    int(x["players"]["0"]['player'][0][1]["player_id"])
                    )
                
                roster_index = strip_team_key(x["players"]["0"]['player'][1]['transaction_data'][0]['destination_team_key']) - 1
                roster = rosters[roster_index].roster
                roster.append(added_player)

            elif x["type"] == 'drop':
                dropped_player_id = x["players"]["0"]['player'][0][1]["player_id"]

                for player in roster:
                    if int(player.player_id) == int(dropped_player_id):
                        roster.remove(player)
                        break
    return rosters

# Change roster gets to multithreading, asyncio?
